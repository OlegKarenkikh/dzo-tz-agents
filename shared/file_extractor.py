import base64
import io
import logging
import os
from typing import Any

import openpyxl
import pdfplumber
import xlrd
from docx import Document
from openai import OpenAI

logger = logging.getLogger("file_extractor")

IMAGE_EXTS = {"jpg", "jpeg", "png", "tiff", "tif", "bmp", "gif", "webp"}

_client: OpenAI | None = None
_client_key: str = ""


def _get_client() -> OpenAI:
    """Returns a cached OpenAI client, re-creating it if API key or base URL changed."""
    global _client, _client_key
    current_key = os.getenv("OPENAI_API_KEY", "ollama")
    current_base = os.getenv("OPENAI_API_BASE", "")
    cache_key = f"{current_key}|{current_base}"
    if _client is None or _client_key != cache_key:
        _client = OpenAI(
            api_key=current_key,
            base_url=current_base or None,
        )
        _client_key = cache_key
    return _client


def extract_text_from_attachment(att: dict[str, Any]) -> str:
    """
    Route text extraction by file type.
    Mirrors the Switch node in n8n: PDF / DOCX / DOC / XLSX / XLS / IMAGE / other.
    """
    ext = att.get("ext", "").lower()
    data = att["data"]
    b64 = att["b64"]
    mime = att.get("mime", "application/octet-stream")

    if ext == "pdf":
        # fix #1: pass only raw bytes; conversion to PNG is handled inside _extract_pdf
        return _extract_pdf(data)
    elif ext == "docx":
        return _extract_docx(data, b64, mime)
    elif ext == "doc":
        # fix #4: DOC is a binary format, not an image.
        # GPT-4o Vision cannot process application/msword via data-URL.
        # Full support requires antiword or LibreOffice conversion.
        filename = att.get("filename", "file.doc")
        logger.warning("DOC format not supported for Vision OCR: %s", filename)
        return (
            f"[Unsupported format .doc ('{filename}'). "
            "Please convert to .docx or .pdf and resend.]"
        )
    elif ext in ("xlsx", "xls"):
        return _extract_spreadsheet(data, ext)
    elif ext in IMAGE_EXTS:
        return _ocr_vision(b64, mime, "IMAGE")
    else:
        return f"[Unsupported format '{att.get('filename', '?')}' ({ext})]"


def _extract_pdf(data: bytes) -> str:
    """
    PDF extraction strategy:
    1. pdfplumber  - for text-based PDFs.
    2. pdf2image   - converts each page to PNG then sends to Vision OCR.

    GPT-4o Vision does NOT accept 'application/pdf' via data-URL;
    scanned PDFs must be converted to image format first.
    """
    try:
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            text = "\n".join(p.extract_text() or "" for p in pdf.pages).strip()
        if len(text) >= 50:
            return text
    except Exception as e:
        logger.warning("pdfplumber failed to read PDF: %s", e)

    # Scanned PDF: convert pages to PNG then OCR each page via Vision
    try:
        from pdf2image import convert_from_bytes  # optional dependency

        images = convert_from_bytes(data, dpi=200, fmt="png")
        page_texts = []
        for i, img in enumerate(images, start=1):
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            b64_page = base64.b64encode(buf.getvalue()).decode()
            page_texts.append(_ocr_vision(b64_page, "image/png", f"PDF-scan p.{i}"))
        return "\n\n".join(page_texts)
    except ImportError:
        logger.error(
            "pdf2image is not installed. "
            "Install with: pip install pdf2image (requires poppler in PATH)."
        )
        return "[OCR error: pdf2image not installed. Requires poppler + pip install pdf2image]"
    except Exception as e:
        logger.error("pdf2image conversion error: %s", e)
        return f"[OCR error during PDF conversion: {e}]"


def _extract_docx(data: bytes, b64: str, mime: str) -> str:
    """DOCX: python-docx first; fall back to Vision OCR on failure."""
    try:
        doc = Document(io.BytesIO(data))
        text = "\n".join(p.text for p in doc.paragraphs).strip()
        if len(text) >= 50:
            return text
    except Exception as e:
        logger.warning("python-docx failed to read DOCX: %s", e)
    return _ocr_vision(b64, mime, "DOCX")


def _ocr_vision(b64: str, mime: str, doc_type: str) -> str:
    """OCR via GPT-4o Vision (or compatible model via OPENAI_API_BASE)."""
    try:
        model = os.getenv("MODEL_NAME", "gpt-4o")
        resp = _get_client().chat.completions.create(
            model=model,
            messages=[
                {
                    "role": "system",
                    "content": (
                        f"You are an OCR engine. Extract ALL text from document {doc_type}. "
                        "Preserve structure: headings, tables (markdown), lists. "
                        "Mark unreadable fragments as [unreadable fragment]. "
                        "Reply with extracted text ONLY."
                    ),
                },
                {
                    "role": "user",
                    "content": [{
                        "type": "image_url",
                        "image_url": {"url": f"data:{mime};base64,{b64}", "detail": "high"},
                    }],
                },
            ],
            max_tokens=4096,
            temperature=0.1,
        )
        return resp.choices[0].message.content
    except Exception as e:
        logger.error("Vision OCR error (%s): %s", doc_type, e)
        return f"[OCR error: {e}]"


def _extract_spreadsheet(data: bytes, ext: str) -> str:
    """XLSX/XLS -> Markdown table."""
    try:
        if ext == "xlsx":
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
            sheets = []
            for name in wb.sheetnames:
                ws = wb[name]
                rows = [[str(c.value or "") for c in row] for row in ws.iter_rows()]
                sheets.append(f"### Sheet: {name}\n" + _rows_to_md(rows))
            return "\n\n".join(sheets)
        else:
            wb = xlrd.open_workbook(file_contents=data)
            sheets = []
            for i in range(wb.nsheets):
                ws = wb.sheet_by_index(i)
                rows = [
                    [str(ws.cell_value(r, c)) for c in range(ws.ncols)]
                    for r in range(ws.nrows)
                ]
                sheets.append(f"### Sheet: {ws.name}\n" + _rows_to_md(rows))
            return "\n\n".join(sheets)
    except Exception as e:
        return f"[Table extraction error: {e}]"


def _rows_to_md(rows: list[list[str]]) -> str:
    if not rows:
        return ""
    header = "| " + " | ".join(rows[0]) + " |"
    sep = "| " + " | ".join(["---"] * len(rows[0])) + " |"
    body = "\n".join("| " + " | ".join(r) + " |" for r in rows[1:])
    return f"{header}\n{sep}\n{body}"
